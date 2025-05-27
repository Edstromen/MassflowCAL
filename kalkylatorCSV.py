import os
import streamlit as st
import pandas as pd
import numpy as np
import altair as alt
from openpyxl import load_workbook

# ----- Konfig för sparande till Excel -----
EXCEL_FILE = "Resultat.xlsx"
SHEET_NAME = "Resultat"

def append_df_to_excel(df, filename=EXCEL_FILE, sheet_name=SHEET_NAME):
    """Appendar en DataFrame till ett Excel-blad, skapar fil/blad vid behov."""
    if not os.path.exists(filename):
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        wb = load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(list(df.columns))
        else:
            ws = wb[sheet_name]
        for row in df.itertuples(index=False):
            ws.append(list(row))
        wb.save(filename)

# ----- Layout & Styling -----
st.set_page_config(page_title="CO₂ Mass Flow Calculator v2.10", layout="wide")
st.markdown("""
    <style>
    .block-container { padding:2rem 3rem; }
    .big-title { font-size:2.5rem; font-weight:700; margin-bottom:1rem; text-align:center; }
    .stHeader, .stSubheader { margin-top:2rem; }
    </style>
""", unsafe_allow_html=True)
st.markdown('<div class="big-title">CO₂ Mass Flow Calculator v2.10</div>', unsafe_allow_html=True)

# ----- Input Mode -----
mode = st.radio("Select input method:", ("Manual input", "Upload CSV"))

# ----- Sidebar-filter -----
with st.sidebar:
    st.header("📐 Filter Settings")
    rotor_diameter   = st.number_input("Rotor diameter (mm)",    min_value=1,  value=350, key="diameter")
    rotor_depth      = st.number_input("Rotor depth (mm)",       min_value=1,  value=100, key="depth")
    active_pct       = st.number_input("Active area (%)",         min_value=1,  max_value=100, value=95, key="active")
    sector_deg_proc  = st.number_input("Absorption sector (°)",  min_value=1,  max_value=360, value=270, key="proc_sector")
    sector_deg_regen = st.number_input("Regeneration sector (°)", min_value=1,  max_value=360, value=90, key="regen_sector")

    st.header("🎯 Scoring Settings")
    threshold_delta_co2 = st.number_input("Threshold: Delta CO₂ per meter (ppm/m)", value=10000, key="thresh_delta")
    threshold_derivata  = st.number_input("Threshold: Derivative GX2_CO2 per m² (ppm/10s/m²)", value=500.0, key="thresh_deriv")
    test_start_ppm = st.number_input("Start test at CO₂ > (ppm)", value=600, key="start_ppm")
    test_stop_ppm  = st.number_input("Stop test at CO₂ > (ppm)", value=1500, key="stop_ppm")

    st.header("🏠 Room Volume")
    room_volume_m3   = st.number_input("Room volume for test (m³)", value=10.5, min_value=0.1, key="room_vol")
    interval_s       = st.number_input("Measurement interval (s)", value=60, min_value=1, key="interval_s")
    ppm_to_mg_per_m3 = st.number_input("mg/m³ per ppm (e.g. 1.96 at 25°C)", value=1.96, key="ppm2mg")

# ----- Beräkna areor -----
rotor_m2      = np.pi * (rotor_diameter / 1000)**2 / 4
area_proc_m2  = rotor_m2 * (active_pct / 100) * (sector_deg_proc / 360)
area_regen_m2 = rotor_m2 * (active_pct / 100) * (sector_deg_regen / 360)

# ----- Hjälpfunktioner -----
def calc_density(T):
    return 1.293 * 273.15 / (273.15 + T)

def calc_abs_humidity(T, RH):
    es = 6.112 * np.exp((17.62 * T) / (243.12 + T))
    e  = RH / 100 * es
    w  = 0.622 * e / (1013.25 - e)
    return w * 1000  # g vatten/kg torr luft

# ----- Manuellt läge (ingen sparfunktion) -----
if mode == "Manual input":
    st.header("⚙️ Manual Parameters")
    c1, c2 = st.columns(2)
    with c1:
        st.subheader("Absorption IN/OUT")
        flow_in_proc  = st.number_input("Flow IN (l/s)", 0.0, 500.0, 73.0, key="m_flow_in_proc")
        T_in_proc     = st.number_input("Temp IN (°C)", -20.0, 200.0, 20.0, key="m_T_in_proc")
        RH_in_proc    = st.number_input("RH IN (%)",    0.0, 100.0, 30.0, key="m_RH_in_proc")
        T_out_proc    = st.number_input("Temp OUT (°C)", -20.0, 200.0, 25.0, key="m_T_out_proc")
        RH_out_proc   = st.number_input("RH OUT (%)",    0.0, 100.0, 20.0, key="m_RH_out_proc")
    with c2:
        st.subheader("Regeneration IN/OUT")
        flow_in_reg   = st.number_input("Flow IN (l/s)", 0.0, 500.0, 30.0, key="m_flow_in_reg")
        T_in_reg      = st.number_input("Temp IN (°C)", -20.0, 200.0, 23.0, key="m_T_in_reg")
        RH_in_reg     = st.number_input("RH IN (%)",    0.0, 100.0, 60.0, key="m_RH_in_reg")
        T_out_reg     = st.number_input("Temp OUT (°C)", -20.0, 200.0, 40.0, key="m_T_out_reg")
        RH_out_reg    = st.number_input("RH OUT (%)",    0.0, 100.0, 50.0, key="m_RH_out_reg")

    # Calculations
    rho_in_proc   = calc_density(T_in_proc)
    rho_out_proc  = calc_density(T_out_proc)
    rho_in_reg    = calc_density(T_in_reg)
    rho_out_reg   = calc_density(T_out_reg)

    ah_in_proc    = calc_abs_humidity(T_in_proc, RH_in_proc)
    ah_out_proc   = calc_abs_humidity(T_out_proc, RH_out_proc)
    ah_in_reg     = calc_abs_humidity(T_in_reg,   RH_in_reg)
    ah_out_reg    = calc_abs_humidity(T_out_reg,  RH_out_reg)

    mf_in_proc    = rho_in_proc  * (flow_in_proc / 1000) / area_proc_m2
    mf_out_proc   = mf_in_proc
    mf_in_reg     = rho_in_reg   * (flow_in_reg  / 1000) / area_regen_m2
    mf_out_reg    = mf_in_reg

    vol_out_proc  = mf_out_proc * area_proc_m2 / rho_out_proc * 1000
    vol_out_reg   = mf_out_reg  * area_regen_m2 / rho_out_reg  * 1000

    ct_proc       = (rotor_depth / 1000) / ((flow_in_proc / 1000) / area_proc_m2)
    ct_reg        = (rotor_depth / 1000) / ((flow_in_reg  / 1000) / area_regen_m2)

    with st.expander("📊 Results (Manual)", expanded=True):
        st.markdown("### Absorption")
        st.write(f"• Mass flow IN:    {mf_in_proc:.3f} kg/m²/s")
        st.write(f"• Mass flow OUT:   {mf_out_proc:.3f} kg/m²/s")
        st.write(f"• Humidity IN:     {ah_in_proc:.1f} g/kg")
        st.write(f"• Humidity OUT:    {ah_out_proc:.1f} g/kg")
        st.write(f"• Volumetric flow IN:   {flow_in_proc:.1f} l/s")
        st.write(f"• Volumetric flow OUT:  {vol_out_proc:.1f} l/s")
        st.write(f"• Contact time:    {ct_proc:.2f} s")

        st.markdown("### Regeneration")
        st.write(f"• Mass flow IN:    {mf_in_reg:.3f} kg/m²/s")
        st.write(f"• Mass flow OUT:   {mf_out_reg:.3f} kg/m²/s")
        st.write(f"• Humidity IN:     {ah_in_reg:.1f} g/kg")
        st.write(f"• Humidity OUT:    {ah_out_reg:.1f} g/kg")
        st.write(f"• Volumetric flow IN:   {flow_in_reg:.1f} l/s")
        st.write(f"• Volumetric flow OUT:  {vol_out_reg:.1f} l/s")
        st.write(f"• Contact time:    {ct_reg:.2f} s")

# ----- CSV-läge med “Spara”-knapp -----
else:
    st.header("📂 Upload CSV for automatic calculation")
    uploaded_files = st.file_uploader(
        "Select one or more CSV files",
        type="csv",
        accept_multiple_files=True,
        key="csvup"
    )

    all_results = []
    all_tests   = []

    if uploaded_files:
        for uploaded in uploaded_files:
            df = pd.read_csv(uploaded)
            df.rename(columns={"GX1_Temp": "GX1_TEMP"}, inplace=True)

            # Radvisa beräkningar
            df["rho_in_proc"]  = calc_density(df["GX3_TEMP"])
            df["rho_out_proc"] = calc_density(df["GX4_TEMP"])
            df["rho_in_reg"]   = calc_density(df["GX2_TEMP"])
            df["rho_out_reg"]  = calc_density(df["GX1_TEMP"])

            df["ah_in_proc"]   = calc_abs_humidity(df["GX3_TEMP"], df["GX3_RH"])
            df["ah_out_proc"]  = calc_abs_humidity(df["GX4_TEMP"], df["GX4_RH"])
            df["ah_in_reg"]    = calc_abs_humidity(df["GX2_TEMP"], df["GX2_RH"])
            df["ah_out_reg"]   = calc_abs_humidity(df["GX1_TEMP"], df["GX1_RH"])

            df["flow_in_proc"] = df["FLOW_Q2"]
            df["flow_in_reg"]  = df["FLOW_Q1"]

            df["mf_in_proc"]   = df["rho_in_proc"] * (df["flow_in_proc"] / 1000) / area_proc_m2
            df["mf_out_proc"]  = df["mf_in_proc"]
            df["mf_in_reg"]    = df["rho_in_reg"] * (df["flow_in_reg"]  / 1000) / area_regen_m2
            df["mf_out_reg"]   = df["mf_in_reg"]

            df["vol_out_proc"] = df["mf_out_proc"] * area_proc_m2 / df["rho_out_proc"] * 1000
            df["vol_out_reg"]  = df["mf_out_reg"]  * area_regen_m2 / df["rho_out_reg"]  * 1000

            df["ct_proc"]      = (rotor_depth / 1000) / ((df["flow_in_proc"] / 1000) / area_proc_m2)
            df["ct_reg"]       = (rotor_depth / 1000) / ((df["flow_in_reg"]  / 1000) / area_regen_m2)
            df["water_added_g_h"] = (df["ah_out_reg"] - df["ah_in_reg"]) * (df["rho_in_reg"] * (df["flow_in_reg"] / 1000)) * 3600

            # Poängberäkning inom testperiod
            try:
                start_idx = df[df["GX2_CO2"] > test_start_ppm].index.min()
                end_idx   = df[df["GX2_CO2"] > test_stop_ppm].index.min()
                if start_idx is None or end_idx is None or start_idx >= end_idx:
                    raise ValueError
                df_test = df.loc[start_idx:end_idx].copy()
                df_test["Delta_CO2"]        = df_test["GX1_CO2"] - df_test["GX2_CO2"]
                df_test["Derivata_GX2"]     = df_test["GX2_CO2"].diff().fillna(0)
                rotor_depth_m               = rotor_depth / 1000
                df_test["Delta_CO2_norm"]   = df_test["Delta_CO2"] / rotor_depth_m
                df_test["Derivata_GX2_norm"] = df_test["Derivata_GX2"] / rotor_m2
                avg_delta = df_test["Delta_CO2_norm"].mean()
                avg_deriv = df_test["Derivata_GX2_norm"].mean()
                score_delta = min(100, avg_delta / threshold_delta_co2 * 100)
                score_deriv  = min(100, avg_deriv / threshold_derivata * 100)
                total_score  = round((score_delta + score_deriv) / 2, 1)
            except:
                df_test      = pd.DataFrame()
                score_delta  = score_deriv = total_score = np.nan
                avg_delta    = avg_deriv  = np.nan

            # ======== NY BERÄKNING KAPACITET via derivata ========
            if not df_test.empty:
                df_test["delta_ppm"]        = df_test["GX2_CO2"].diff().fillna(0)
                df_test["delta_mg_per_m3"]  = df_test["delta_ppm"] * ppm_to_mg_per_m3
                df_test["uptake_mg"]        = df_test["delta_mg_per_m3"] * room_volume_m3
                total_mg                    = df_test["uptake_mg"].sum()
                test_time_h                 = len(df_test) * interval_s / 3600
                regen_capacity_kg_24h       = ((total_mg / 1e6) / test_time_h * 24) if test_time_h > 0 else np.nan
            else:
                regen_capacity_kg_24h = np.nan
            # ======== SLUT ny beräkning ========

            # Aggrera medelvärden och poäng
            res = {
                "Abs IN mf (kg/m²/s)": df["mf_in_proc"].mean(),
                "Reg IN mf (kg/m²/s)": df["mf_in_reg"].mean(),
                "Diff mf (kg/m²/s)": df["mf_in_proc"].mean() - df["mf_in_reg"].mean(),
                "Abs IN vol (l/s)": df["flow_in_proc"].mean(),
                "Abs UT vol (l/s)": df["vol_out_proc"].mean(),
                "Reg IN vol (l/s)": df["flow_in_reg"].mean(),
                "Reg UT vol (l/s)": df["vol_out_reg"].mean(),
                "Abs IN ah (g/kg)": df["ah_in_proc"].mean(),
                "Abs UT ah (g/kg)": df["ah_out_proc"].mean(),
                "Reg IN ah (g/kg)": df["ah_in_reg"].mean(),
                "Reg UT ah (g/kg)": df["ah_out_reg"].mean(),
                "Kontakttid Abs (s)": df["ct_proc"].mean(),
                "Kontakttid Reg (s)": df["ct_reg"].mean(),
                "Vatten tillsatt (g/h)": df["water_added_g_h"].mean(),
                "Poäng ΔCO₂": score_delta,
                "Poäng derivata": score_deriv,
                "Total poäng": total_score,
                "Testpunkter": len(df_test),
                "ΔCO₂ (medel ppm/m)": avg_delta,
                "Derivata (medel ppm/10s/m²)": avg_deriv,
                "CO₂-kapacitet (kg/24h)": regen_capacity_kg_24h,
            }
            # --- Calculate mean derivative in ppm per 30min ---
            if not df_test.empty and len(df_test) > 1:
                # Mean derivative per second
                mean_deriv_per_sec = df_test["Derivata_GX2"].mean() / interval_s
                deriv_ppm_per_30min = mean_deriv_per_sec * 1800  # 30min = 1800s
            else:
                deriv_ppm_per_30min = np.nan
            res["Derivata (ppm/30min)"] = deriv_ppm_per_30min

            df_res = pd.Series(res, name="Mean Value").to_frame().T.reset_index(drop=True)
            df_res["Mode"]       = "CSV"
            df_res["SourceFile"] = uploaded.name
            all_results.append(df_res)

            # Spara testperiod för tidsseriegraf
            if not df_test.empty:
                df_plot = df_test.reset_index().copy()
                df_plot["rel_index"]  = df_plot.index
                df_plot["SourceFile"] = uploaded.name
                all_tests.append(df_plot)

    # Sammanställning efter loopen
    if all_results:
        combined_df = pd.concat(all_results, ignore_index=True)
        st.subheader("📋 File Comparison")
        st.dataframe(combined_df, use_container_width=True)

       

        # GX2_CO₂ over the test period
        if all_tests:
            st.subheader("📈 GX2_CO₂ Comparison Over Test Period")
            ts_df = pd.concat(all_tests, ignore_index=True)
            ts_chart = (
                alt.Chart(ts_df)
                   .mark_line(point=False)
                   .encode(
                       x=alt.X("rel_index:Q", title=f"Time index since test start ({interval_s}s interval)"),
                       y=alt.Y("GX2_CO2:Q", title="CO₂ (ppm)"),
                       color=alt.Color("SourceFile:N", title="File"),
                       tooltip=["SourceFile", "index", "GX2_CO2"]
                   )
                   .properties(width=700, height=400)
            )
            st.altair_chart(ts_chart, use_container_width=True)

            # --- Standardized Bar Charts: All after the main graph ---


            # 2. CO₂ capacity bar chart
            st.subheader("🧲 CO₂ Capacity (kg/24h) per File")
            if "CO₂-kapacitet (kg/24h)" in combined_df.columns:
                cap_df = combined_df[["SourceFile", "CO₂-kapacitet (kg/24h)"]]
                cap_chart = (
                    alt.Chart(cap_df)
                       .mark_bar(size=60)
                       .encode(
                           x=alt.X("SourceFile:N", title="Filename"),
                           y=alt.Y("CO₂-kapacitet (kg/24h):Q", title="CO₂ Capacity (kg/24h)"),
                           tooltip=["SourceFile", "CO₂-kapacitet (kg/24h)"],
                           color=alt.Color("SourceFile:N", legend=None)
                       )
                       .properties(width=600, height=300)
                )
                st.altair_chart(cap_chart, use_container_width=False)
            else:
                st.warning("No CO₂ capacity data to display.")

            # 3. Derivative (ppm/30min) bar chart
            st.subheader("📈 Derivative (ppm/30min) per File")
            if "Derivata (ppm/30min)" in combined_df.columns:
                deriv30_df = combined_df[["SourceFile", "Derivata (ppm/30min)"]]
                deriv30_chart = (
                    alt.Chart(deriv30_df)
                       .mark_bar(size=60)
                       .encode(
                           x=alt.X("SourceFile:N", title="Filename"),
                           y=alt.Y("Derivata (ppm/30min):Q", title="Derivative (ppm/30min)"),
                           tooltip=["SourceFile", "Derivata (ppm/30min)"],
                           color=alt.Color("SourceFile:N", legend=None)
                       )
                       .properties(width=600, height=300)
                )
                st.altair_chart(deriv30_chart, use_container_width=False)
            else:
                st.warning("No derivative (ppm/30min) data to display.")

        # (Duplicated CO₂-kapacitet stapeldiagram removed)
